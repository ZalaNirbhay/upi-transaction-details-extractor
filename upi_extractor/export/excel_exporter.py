"""
Excel Exporter Module — Enhanced
---------------------------------
Exports extracted payment data to a professionally formatted Excel file.
Features:
  • Styled header row with colored background
  • Auto-fit column widths
  • Conditional formatting (green credits, red debits)
  • Summary row with totals for amount columns
  • Handles both UPI and Passbook column layouts
"""

import os
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Style Constants ──────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

DATA_FONT = Font(name="Calibri", size=10)
DATA_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=False)
AMOUNT_ALIGN = Alignment(horizontal="right", vertical="center")

CREDIT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
CREDIT_FONT = Font(name="Calibri", size=10, color="006100")
DEBIT_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
DEBIT_FONT = Font(name="Calibri", size=10, color="9C0006")

SUMMARY_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUMMARY_FONT = Font(name="Calibri", bold=True, size=11, color="1F4E79")

THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

# Amount-related column names (used for summation + formatting)
AMOUNT_COLUMNS = {
    'Amount', 'Credit (₹)', 'Debit (₹)', 'Balance (₹)',
    'Opening Balance (₹)', 'Closing Balance (₹)',
}


def _auto_fit_columns(ws, df):
    """Adjust column widths to fit the longest value in each column."""
    for col_idx, col_name in enumerate(df.columns, start=1):
        # Measure header width
        max_width = len(str(col_name)) + 4

        # Measure data widths (sample first 100 rows for performance)
        for row in df.head(100).itertuples(index=False):
            cell_value = str(row[col_idx - 1]) if row[col_idx - 1] else ''
            max_width = max(max_width, len(cell_value) + 2)

        # Cap to reasonable maximum
        max_width = min(max_width, 45)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_width


def _style_header(ws, num_cols):
    """Apply professional styling to the header row."""
    ws.row_dimensions[1].height = 30
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _style_data_rows(ws, df, num_rows, num_cols):
    """Apply formatting to data cells with special treatment for amounts."""
    amount_col_indices = set()
    for col_idx, col_name in enumerate(df.columns, start=1):
        if col_name in AMOUNT_COLUMNS:
            amount_col_indices.add(col_idx)

    for row in range(2, num_rows + 2):  # +2 because header is row 1
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER

            if col in amount_col_indices:
                cell.alignment = AMOUNT_ALIGN
                # Try to set number format for amount cells
                try:
                    val = float(str(cell.value).replace(',', ''))
                    cell.value = val
                    cell.number_format = '#,##0.00'
                except (ValueError, TypeError):
                    pass
            else:
                cell.alignment = DATA_ALIGN

    # Alternate row shading for readability
    light_gray = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    for row in range(3, num_rows + 2, 2):  # Every other row starting from row 3
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            if cell.fill == PatternFill():  # Only if not already filled
                cell.fill = light_gray


def _apply_conditional_formatting(ws, df, num_rows):
    """Color credit cells green and debit cells red."""
    for col_idx, col_name in enumerate(df.columns, start=1):
        if 'credit' in col_name.lower() or col_name == 'Amount':
            for row in range(2, num_rows + 2):
                cell = ws.cell(row=row, column=col_idx)
                if cell.value and str(cell.value).strip():
                    cell.fill = CREDIT_FILL
                    cell.font = CREDIT_FONT

        elif 'debit' in col_name.lower():
            for row in range(2, num_rows + 2):
                cell = ws.cell(row=row, column=col_idx)
                if cell.value and str(cell.value).strip():
                    cell.fill = DEBIT_FILL
                    cell.font = DEBIT_FONT


def _add_summary_row(ws, df, num_rows, num_cols):
    """Add a totals row at the bottom for amount columns."""
    summary_row = num_rows + 2  # +2 because data starts from row 2

    # Label
    label_cell = ws.cell(row=summary_row, column=1)
    label_cell.value = "TOTALS"
    label_cell.font = SUMMARY_FONT
    label_cell.fill = SUMMARY_FILL
    label_cell.border = THIN_BORDER
    label_cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=summary_row, column=col_idx)
        cell.fill = SUMMARY_FILL
        cell.border = THIN_BORDER
        cell.font = SUMMARY_FONT

        if col_name in AMOUNT_COLUMNS:
            # Sum the column values
            total = 0.0
            has_values = False
            for data_row in range(2, num_rows + 2):
                val = ws.cell(row=data_row, column=col_idx).value
                try:
                    total += float(str(val).replace(',', ''))
                    has_values = True
                except (ValueError, TypeError):
                    pass

            if has_values:
                cell.value = total
                cell.number_format = '#,##0.00'
                cell.alignment = AMOUNT_ALIGN


def export_to_excel(data_list, output_path, append=False):
    """
    Save extracted payment data to a professionally formatted Excel file.

    Features:
        - Styled header with dark-blue background
        - Auto-fit column widths
        - Amount columns right-aligned with number format
        - Credits highlighted green, debits highlighted red
        - Summary row with totals at bottom
        - Alternate row shading for readability
        - Append mode: add data to existing file without overwriting

    Args:
        data_list (list[dict]): List of extracted payment records.
        output_path (str): Destination file path (must end with .xlsx).
        append (bool): If True, append to existing file (default: False).

    Returns:
        tuple: (success: bool, message: str)
    """
    if not data_list:
        return False, "No data to export."

    # Ensure .xlsx extension
    if not output_path.lower().endswith('.xlsx'):
        output_path += '.xlsx'

    try:
        df_new = pd.DataFrame(data_list)

        # Drop "All Extracted Text" column — too long for Excel readability
        if 'All Extracted Text' in df_new.columns:
            df_new = df_new.drop(columns=['All Extracted Text'])

        # ── Append mode: load existing data and concatenate ──
        if append and os.path.exists(output_path):
            try:
                df_existing = pd.read_excel(output_path, engine='openpyxl')
                # Drop the summary/totals row if present (artifact of formatting)
                first_col = df_existing.columns[0]
                df_existing = df_existing[df_existing[first_col] != 'TOTALS']
                df = pd.concat([df_existing, df_new], ignore_index=True)
                mode_label = f"Appended {len(df_new)} row(s) to"
            except Exception:
                df = df_new
                mode_label = "Successfully saved to"
        else:
            df = df_new
            mode_label = "Successfully saved to"

        num_rows = len(df)
        num_cols = len(df.columns)

        # Write base data using pandas (creates the workbook)
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Extracted Data')
            ws = writer.sheets['Extracted Data']

            # Apply formatting
            _style_header(ws, num_cols)
            _style_data_rows(ws, df, num_rows, num_cols)
            _apply_conditional_formatting(ws, df, num_rows)
            _add_summary_row(ws, df, num_rows, num_cols)
            _auto_fit_columns(ws, df)

            # Freeze top row for scrolling
            ws.freeze_panes = 'A2'

        return True, f"{mode_label} {output_path} ({num_rows} total rows)"

    except Exception as e:
        return False, f"Error saving Excel: {e}"

